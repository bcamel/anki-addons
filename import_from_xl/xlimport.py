# import the main window object (mw) from ankiqt
from aqt import mw

# import the "show info" tool from utils.py
# from aqt.utils import showInfo
# import all of the Qt GUI library
from aqt.qt import *

import re
import anki.utils
import anki.importing
import anki.hooks
from anki.importing.noteimp import NoteImporter

from openpyxl import load_workbook
import pprint as pp
#from aqt.qt import debug; debug()

class ExcelImporter(NoteImporter):

    class _Data(object):
        def __init__(self, log):
             self.worksheet = None
             self.header = None
             self._headerCells = None
             self._log = log
             self._numFields = 0
             self._dataDim = None
             self._rowOffset = 0
             self._notes = None

        def hasOpenWorkbook(self):
            return self.worksheet != None

        def openWorkbook(self, file):
            if self.hasOpenWorkbook():
                return
            try:
                workbook = load_workbook(filename = file)
            except Exception, e:
                msg = repr(str(e))
                raise IOError(u"<%s> not a valid Excel file: %s" % (file, msg))
            for sheet in workbook:
                self._log.append("openWorkbook: checking '%s'" % sheet.title)
                if self.checkValidSheet(sheet):
                    self.worksheet = sheet
                    # TODO currently just pick the first one
                    break

        def checkValidSheet(self, sheet):
            dim = sheet.calculate_dimension()
            # TODO think about how to deal with sheets w/o header
            header = None
            self._rowOffset = 0
            for row in tuple(sheet.iter_rows(dim)):
                self._rowOffset += 1
                vals = len(filter((lambda y: y.value), row))
                if not header and vals > 0:
                    header = row
                elif len(filter((lambda y: y.value), header)) < vals:
                    header = row
                    break        
            if header:
                self.header = map((lambda x: str(x.value)), filter((lambda y: y.value), header))
                self._headerCells = header
                self._numFields = len(filter((lambda y: y.value), header))
                self._dataDim = dim
                return True
            else:
                return False

        def readData(self, mapping):
            from anki.importing.noteimp import ForeignNote
            if not self.hasOpenWorkbook():
                self._log.append("xlimport: Error reading data: no open worksheet.")
                return
            assert self._dataDim
            assert self._headerCells
            self._notes = []
            for row in tuple(self.worksheet.iter_rows(self._dataDim, row_offset = self._rowOffset)):
                flds = []
                isEmpty = True
                for c in xrange(0, len(mapping)):
                    hdr = self._headerCells[c].value
                    if hdr and hdr in mapping:
                        v = row[c].value if row[c].value else ''
                        isEmpty = False if row[c].value else isEmpty
                        flds.append(v if row[c].data_type == 's' else str(v))
                if not isEmpty:
                    note = ForeignNote()
                    note.fields = flds
                    self._notes.append(note)


    needMapper = True
    
    def __init__(self, *args):
        NoteImporter.__init__(self, *args)
        self._data = None

    def _loadWorkbook(self):
        if not self._data:
            self._data = self._Data(self.log)
            self._data.openWorkbook(self.file)

    def _chooseLikelyModel(self):
        likelyModel = None
        xlFieldsInModel = 0
        def isOcc(x):
            return x in self._data.header
        for m in self.col.models.all():
            fieldNames = [f['name'] for f in m['flds']]
            occ = len(filter((lambda x : x), [isOcc(f) for f in fieldNames]))
            if xlFieldsInModel < occ:
                likelyModel = m['name']
                xlFieldsInModel = occ
        # TODO: borrowed from ModelChooser.onModelChange (better ways to update?)
        self.model = self.col.models.byName(likelyModel)
        self.col.conf['curModel'] = self.model['id']
        cdeck = self.col.decks.current()
        cdeck['mid'] = self.model['id']
        self.col.decks.save(cdeck)
        anki.hooks.runHook("currentModelChanged")
        mw.reset()

    def open(self):
        self._loadWorkbook()
        if not self._data.hasOpenWorkbook():
            if self._data.worksheet:
                raise IOError(u"Could not load data from worksheet %s" % self.worksheet.title)
            else:
                raise IOError(u"Could not load data from file %s" % self.file)

    def fields(self):
        "Number of fields."
        self.open()
        return self._data._numFields

    def foreignNotes(self):
        "Import."
        assert self.mapping
        self._data.readData(tuple(self.mapping))
        return self._data._notes

    def initMapping(self):
        self._chooseLikelyModel()
        NoteImporter.initMapping(self)
        # let the mappings turn out in the order of the spreadsheet headers
        mapping = self.mapping[0:]
        for i in xrange(0, len(self._data.header)):
            try:
                h = self._data.header[i]
                j = mapping.index(h)
                self.mapping[i] = mapping[j]
            except ValueError:
                self.log.append("Unable to find mapping for header '%s'" % h)
                self.mapping[i] = None
        #pp.pprint(mapping)
        #pp.pprint(self.mapping)





anki.importing.Importers =  (
    ((u"Excel 2007/2010/2013 Worksheet (*.xlsx)"), ExcelImporter),)



# pp.pprint(importing.Importers)
#importing.Importers = importing.Importers + (
#    ((u"Excel 2007/2010/2013 Worksheet (*.xlsx)"), ExcelImporter),)
# print
# pp.pprint(importing.Importers)